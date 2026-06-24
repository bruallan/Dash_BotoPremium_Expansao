import os
import time
import uuid
from playwright.sync_api import sync_playwright
from supabase import create_client, Client
from dotenv import load_dotenv

# Carrega as variáveis de ambiente do arquivo .env (caso executado localmente)
load_dotenv()

# ==========================================
# 1. Configurações Iniciais e Credenciais
# ==========================================
url: str = os.environ.get("VITE_SUPABASE_URL")
key: str = os.environ.get("VITE_SUPABASE_SERVICE_ROLE_KEY")
sults_user: str = os.environ.get("SULTS_USER")
sults_pass: str = os.environ.get("SULTS_PASS")

supabase: Client = create_client(url, key)

def main():
    print("Iniciando rotina de sincronização...")

    # ==========================================
    # 2. Mapeamento do Banco de Dados Atual
    # ==========================================
    print("Consultando base atual de colaboradores no Supabase...")
    response = supabase.table("colaboradores").select("id, email, status").execute()
    
    usuarios_banco = {
        linha["email"]: {"id": linha["id"], "status": linha.get("status", "")} 
        for linha in response.data
    }
    
    emails_encontrados_no_portal = set()

    # ==========================================
    # 3. Navegação Web (Playwright)
    # ==========================================
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        print("Acessando o portal...")
        page.goto("https://botopremium.sults.com.br/")
        
        # Preenchendo os campos com os IDs exatos que você mapeou!
        page.fill("input[id='form:login-usuario-inputText']", sults_user) 
        page.fill("input[id='form:login-user-password']", sults_pass)   
        
        # Truque: Simulando a tecla "Enter" no campo de senha para fazer login
        try:
            with page.expect_navigation(timeout=30000):
                page.press("input[id='form:login-user-password']", "Enter")
        except Exception as e:
            print(f"Aviso ao aguardar navegação pós-login: {e}")
            print(f"URL atual: {page.url}")
            try:
                print("Texto da tela atual:")
                print(page.evaluate("document.body.innerText")[:1500])
            except:
                pass
        
        # ==========================================
        # 3. Exportar Planilha
        # ==========================================
        print("Acessando página de exportação...")
        try:
            page.goto("https://botopremium.sults.com.br/empresa/exportar", wait_until="domcontentloaded", timeout=60000)
        except Exception as e:
            print(f"Erro no goto exportar: {e}")

        print("Clicando no botão Exportar Colaboradores...")
        # A div do card com ícone mdi-account
        page.wait_for_selector("div.card-export[onclick*='exportPessoaDialogWVar']", timeout=15000)
        page.locator("div.card-export[onclick*='exportPessoaDialogWVar']").click()
        
        print("Aguardando download da planilha...")
        with page.expect_download(timeout=120000) as download_info:
            page.wait_for_selector("button[onclick*='PrimeFaces.monitorDownload']", timeout=15000)
            page.locator("button[onclick*='PrimeFaces.monitorDownload']").click()
            
        download = download_info.value
        download_path = download.path()
        print(f"Download concluído: {download_path}")

        browser.close()

        # ==========================================
        # 4. Ler Planilha Excel
        # ==========================================
        import openpyxl
        wb = openpyxl.load_workbook(download_path)
        ws = wb.active

        header = [cell.value for cell in ws[1]]
        
        idx_id = header.index("ID") if "ID" in header else None
        idx_nome = header.index("Nome") if "Nome" in header else None
        idx_telefone = header.index("Telefones") if "Telefones" in header else None
        idx_email = header.index("Email") if "Email" in header else None
        idx_cargo = header.index("Cargo") if "Cargo" in header else None
        
        if any(idx is None for idx in [idx_id, idx_nome, idx_telefone, idx_email, idx_cargo]):
            print("Erro: Não foi possível encontrar todas as colunas necessárias na planilha.")
            print("Cabeçalhos encontrados:", header)
            return

        emails_encontrados_no_portal = set()

        # ==========================================
        # 5. Lógica Inteligente de Inserir ou Atualizar
        # ==========================================
        print("Iniciando leitura dos colaboradores na planilha...")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx_nome] is None and row[idx_email] is None:
                continue # Linha vazia
                
            colab_id = str(row[idx_id]).strip() if row[idx_id] is not None else str(uuid.uuid4())
            nome = str(row[idx_nome]).strip() if row[idx_nome] else ""
            celular = str(row[idx_telefone]).strip() if row[idx_telefone] else ""
            email = str(row[idx_email]).strip() if row[idx_email] else ""
            cargo = str(row[idx_cargo]).strip() if row[idx_cargo] else ""

            if not email or "@" not in email:
                print(f"Ignorando {nome} - Sem email válido ({email})")
                continue

            print(f"Lido: {nome} | {email} | {cargo} | {celular} | ID: {colab_id}")
            
            emails_encontrados_no_portal.add(email)

            if email in usuarios_banco:
                id_existente = usuarios_banco[email]["id"]
                dados_atualizacao = {
                    "id": colab_id, # Caso o banco permita atualizar a PK
                    "nome_colaborador": nome,
                    "celular": celular,
                    "cargo": cargo, 
                    "status": "ativo"
                }
                try:
                    supabase.table("colaboradores").update(dados_atualizacao).eq("id", id_existente).execute()
                except Exception as e:
                    # Se falhar porque id não pode ser alterado, tenta atualizar sem alterar o id
                    dados_atualizacao.pop("id")
                    supabase.table("colaboradores").update(dados_atualizacao).eq("id", id_existente).execute()
            else:
                dados_insercao = {
                    "id": colab_id,
                    "nome_colaborador": nome,
                    "celular": celular,
                    "email": email,
                    "cargo": cargo, 
                    "status": "ativo"
                }
                try:
                    supabase.table("colaboradores").insert(dados_insercao).execute()
                    usuarios_banco[email] = {"id": colab_id, "status": "ativo"}
                except Exception as e:
                    print(f"Erro ao inserir {nome}: {e}")

    # ==========================================
    # 7. Atualizar status para Inativo
    # ==========================================
    print("Verificando colaboradores inativos...")
    for email, info in usuarios_banco.items():
        if info["status"] == "ativo" and email not in emails_encontrados_no_portal:
            supabase.table("colaboradores").update({"status": "inativo"}).eq("id", info["id"]).execute()
            print(f"Colaborador inativado: {email}")

    print("Rotina finalizada com sucesso!")

if __name__ == "__main__":
    main()